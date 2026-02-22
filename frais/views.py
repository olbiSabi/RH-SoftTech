# frais/views.py
"""
Vues pour le module Notes de Frais.
"""
from django.conf import settings
from django.core.cache import cache
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.views.decorators.http import require_POST, require_GET
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.utils import timezone
from datetime import datetime

_CACHE_TTL = getattr(settings, 'CACHE_TTL_STATS', 3600)
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from frais.models import NFNF, NFLF, NFAV, NFCA
from frais.forms import (
    NoteFraisForm, LigneFraisForm, AvanceForm,
    ApprobationAvanceForm, RejetForm, RemboursementForm,
    VersementAvanceForm, FiltreNotesForm, FiltreAvancesForm,
    ValidationLigneForm
)
from frais.services import (
    NoteFraisService, AvanceService, CategorieService,
    ValidationFraisService, StatistiquesFraisService
)


# =============================================================================
# DASHBOARD
# =============================================================================

@login_required
def dashboard_frais(request):
    """Dashboard principal du module frais."""
    employe = request.user.employe

    # Stats personnelles
    mes_stats = StatistiquesFraisService.get_stats_employe(employe)

    # Mes notes récentes
    mes_notes = NoteFraisService.get_notes_employe(employe)[:5]

    # Mes avances en cours
    mes_avances = AvanceService.get_avances_employe(employe).exclude(
        STATUT__in=['REGULARISE', 'ANNULE']
    )[:5]

    context = {
        'stats': mes_stats,
        'mes_notes': mes_notes,
        'mes_avances': mes_avances,
    }

    # Stats globales pour les valideurs
    if _peut_valider(employe):
        context['notes_a_valider'] = NoteFraisService.get_notes_a_valider().count()
        context['avances_a_approuver'] = AvanceService.get_avances_a_approuver().count()
        context['stats_globales'] = StatistiquesFraisService.get_stats_globales()

    return render(request, 'frais/dashboard.html', context)


# =============================================================================
# NOTES DE FRAIS - CRUD
# =============================================================================

@login_required
def liste_notes_frais(request):
    """Liste des notes de frais de l'employé connecté."""
    employe = request.user.employe
    form = FiltreNotesForm(request.GET)

    notes = NoteFraisService.get_notes_employe(employe)

    # Appliquer les filtres
    if form.is_valid():
        if form.cleaned_data.get('statut'):
            notes = notes.filter(STATUT=form.cleaned_data['statut'])
        if form.cleaned_data.get('date_debut'):
            notes = notes.filter(PERIODE_DEBUT__gte=form.cleaned_data['date_debut'])
        if form.cleaned_data.get('date_fin'):
            notes = notes.filter(PERIODE_FIN__lte=form.cleaned_data['date_fin'])

    # Pagination
    paginator = Paginator(notes, 10)
    page = request.GET.get('page', 1)
    notes_page = paginator.get_page(page)

    return render(request, 'frais/notes/liste.html', {
        'notes': notes_page,
        'form': form,
    })


@login_required
def creer_note_frais(request):
    """Création d'une nouvelle note de frais."""
    employe = request.user.employe

    if request.method == 'POST':
        form = NoteFraisForm(request.POST)
        if form.is_valid():
            try:
                note = NoteFraisService.creer_note_frais(
                    employe=employe,
                    periode_debut=form.cleaned_data['PERIODE_DEBUT'],
                    periode_fin=form.cleaned_data['PERIODE_FIN'],
                    objet=form.cleaned_data.get('OBJET'),
                    created_by=employe
                )
                messages.success(request, f"Note de frais {note.REFERENCE} créée avec succès")
                return redirect('frais:detail_note', uuid=note.uuid)
            except Exception as e:
                messages.error(request, str(e))
    else:
        form = NoteFraisForm()

    return render(request, 'frais/notes/creer.html', {'form': form})


@login_required
def detail_note_frais(request, uuid):
    """Détail d'une note de frais."""
    note = get_object_or_404(NFNF, uuid=uuid)
    employe = request.user.employe

    # Vérifier l'accès
    if note.EMPLOYE != employe and not _peut_valider(employe):
        return HttpResponseForbidden("Accès non autorisé")

    # Formulaire d'ajout de ligne
    ligne_form = LigneFraisForm(note_frais=note)

    # Anomalies détectées
    anomalies = ValidationFraisService.get_anomalies_note(note)

    context = {
        'note': note,
        'lignes': note.lignes.select_related('CATEGORIE').all(),
        'ligne_form': ligne_form,
        'anomalies': anomalies,
        'peut_modifier': note.peut_etre_modifie() and note.EMPLOYE == employe,
        'peut_valider': _peut_valider(employe) and note.peut_etre_valide(),
    }

    return render(request, 'frais/notes/detail.html', context)


@login_required
def modifier_note_frais(request, uuid):
    """Modification d'une note de frais."""
    note = get_object_or_404(NFNF, uuid=uuid)
    employe = request.user.employe

    if note.EMPLOYE != employe:
        return HttpResponseForbidden("Accès non autorisé")

    if not note.peut_etre_modifie():
        messages.error(request, "Cette note ne peut plus être modifiée")
        return redirect('frais:detail_note', uuid=uuid)

    if request.method == 'POST':
        form = NoteFraisForm(request.POST, instance=note)
        if form.is_valid():
            try:
                NoteFraisService.modifier_note_frais(
                    note,
                    PERIODE_DEBUT=form.cleaned_data['PERIODE_DEBUT'],
                    PERIODE_FIN=form.cleaned_data['PERIODE_FIN'],
                    OBJET=form.cleaned_data.get('OBJET')
                )
                messages.success(request, "Note de frais modifiée")
                return redirect('frais:detail_note', uuid=uuid)
            except Exception as e:
                messages.error(request, str(e))
    else:
        form = NoteFraisForm(instance=note)

    return render(request, 'frais/notes/modifier.html', {
        'form': form,
        'note': note
    })


@login_required
@require_POST
def supprimer_note_frais(request, uuid):
    """Suppression d'une note de frais (brouillon uniquement)."""
    note = get_object_or_404(NFNF, uuid=uuid)
    employe = request.user.employe

    if note.EMPLOYE != employe:
        return HttpResponseForbidden("Accès non autorisé")

    if note.STATUT != 'BROUILLON':
        messages.error(request, "Seules les notes en brouillon peuvent être supprimées")
        return redirect('frais:detail_note', uuid=uuid)

    reference = note.REFERENCE
    note.delete()
    messages.success(request, f"Note {reference} supprimée")
    return redirect('frais:liste_notes')


# =============================================================================
# LIGNES DE FRAIS
# =============================================================================

@login_required
@require_POST
def ajouter_ligne(request, note_uuid):
    """Ajoute une ligne à une note de frais."""
    note = get_object_or_404(NFNF, uuid=note_uuid)
    employe = request.user.employe

    if note.EMPLOYE != employe:
        return JsonResponse({'success': False, 'error': 'Accès non autorisé'}, status=403)

    if not note.peut_etre_modifie():
        return JsonResponse({'success': False, 'error': 'Note non modifiable'}, status=400)

    form = LigneFraisForm(request.POST, request.FILES, note_frais=note)

    if form.is_valid():
        try:
            ligne = NoteFraisService.ajouter_ligne(
                note=note,
                categorie=form.cleaned_data['CATEGORIE'],
                date_depense=form.cleaned_data['DATE_DEPENSE'],
                description=form.cleaned_data['DESCRIPTION'],
                montant=form.cleaned_data['MONTANT'],
                justificatif=form.cleaned_data.get('JUSTIFICATIF'),
                numero_facture=form.cleaned_data.get('NUMERO_FACTURE'),
                devise=form.cleaned_data.get('DEVISE', 'XOF')
            )
            return JsonResponse({
                'success': True,
                'ligne_id': ligne.id,
                'montant_total': str(note.MONTANT_TOTAL)
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    else:
        errors = {field: errors[0] for field, errors in form.errors.items()}
        return JsonResponse({'success': False, 'errors': errors}, status=400)


@login_required
@require_POST
def supprimer_ligne(request, ligne_uuid):
    """Supprime une ligne de frais."""
    ligne = get_object_or_404(NFLF, uuid=ligne_uuid)
    employe = request.user.employe

    if ligne.NOTE_FRAIS.EMPLOYE != employe:
        return JsonResponse({'success': False, 'error': 'Accès non autorisé'}, status=403)

    try:
        note = ligne.NOTE_FRAIS
        NoteFraisService.supprimer_ligne(ligne)
        return JsonResponse({
            'success': True,
            'montant_total': str(note.MONTANT_TOTAL)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


# =============================================================================
# WORKFLOW NOTES DE FRAIS
# =============================================================================

@login_required
@require_POST
def soumettre_note(request, uuid):
    """Soumet une note de frais pour validation."""
    note = get_object_or_404(NFNF, uuid=uuid)
    employe = request.user.employe

    if note.EMPLOYE != employe:
        return HttpResponseForbidden("Accès non autorisé")

    try:
        # Valider la note avant soumission
        validation = ValidationFraisService.valider_note_frais(note)
        if not validation['is_valid']:
            for error in validation['errors']:
                messages.error(request, error)
            return redirect('frais:detail_note', uuid=uuid)

        NoteFraisService.soumettre_note(note)
        messages.success(request, f"Note {note.REFERENCE} soumise pour validation")
    except Exception as e:
        messages.error(request, str(e))

    return redirect('frais:detail_note', uuid=uuid)


@login_required
@require_POST
def valider_note(request, uuid):
    """Valide une note de frais."""
    note = get_object_or_404(NFNF, uuid=uuid)
    employe = request.user.employe

    if not _peut_valider(employe):
        return HttpResponseForbidden("Accès non autorisé")

    try:
        commentaire = request.POST.get('commentaire', '')
        NoteFraisService.valider_note(note, employe, commentaire)
        messages.success(request, f"Note {note.REFERENCE} validée")
    except Exception as e:
        messages.error(request, str(e))

    return redirect('frais:detail_note', uuid=uuid)


@login_required
@require_POST
def rejeter_note(request, uuid):
    """Rejette une note de frais."""
    note = get_object_or_404(NFNF, uuid=uuid)
    employe = request.user.employe

    if not _peut_valider(employe):
        return HttpResponseForbidden("Accès non autorisé")

    form = RejetForm(request.POST)
    if form.is_valid():
        try:
            NoteFraisService.rejeter_note(
                note, employe, form.cleaned_data['commentaire']
            )
            messages.success(request, f"Note {note.REFERENCE} rejetée")
        except Exception as e:
            messages.error(request, str(e))
    else:
        messages.error(request, "Un motif de rejet est obligatoire")

    return redirect('frais:detail_note', uuid=uuid)


@login_required
@require_POST
def rembourser_note(request, uuid):
    """Marque une note comme remboursée."""
    note = get_object_or_404(NFNF, uuid=uuid)
    employe = request.user.employe

    if not _peut_valider(employe):
        return HttpResponseForbidden("Accès non autorisé")

    form = RemboursementForm(request.POST)
    if form.is_valid():
        try:
            NoteFraisService.marquer_rembourse(
                note,
                form.cleaned_data['date_remboursement'],
                form.cleaned_data.get('reference_paiement')
            )
            messages.success(request, f"Note {note.REFERENCE} marquée comme remboursée")
        except Exception as e:
            messages.error(request, str(e))
    else:
        messages.error(request, "Veuillez renseigner la date de remboursement")

    return redirect('frais:detail_note', uuid=uuid)


# =============================================================================
# AVANCES SUR FRAIS
# =============================================================================

@login_required
def liste_avances(request):
    """Liste des avances de l'employé connecté."""
    employe = request.user.employe
    form = FiltreAvancesForm(request.GET)

    avances = AvanceService.get_avances_employe(employe)

    if form.is_valid():
        if form.cleaned_data.get('statut'):
            avances = avances.filter(STATUT=form.cleaned_data['statut'])

    paginator = Paginator(avances, 10)
    page = request.GET.get('page', 1)
    avances_page = paginator.get_page(page)

    return render(request, 'frais/avances/liste.html', {
        'avances': avances_page,
        'form': form,
        'solde_en_cours': AvanceService.get_solde_avances_employe(employe)
    })


@login_required
def creer_avance(request):
    """Création d'une demande d'avance."""
    employe = request.user.employe

    if request.method == 'POST':
        form = AvanceForm(request.POST)
        if form.is_valid():
            try:
                avance = AvanceService.creer_avance(
                    employe=employe,
                    montant_demande=form.cleaned_data['MONTANT_DEMANDE'],
                    motif=form.cleaned_data['MOTIF'],
                    date_mission_debut=form.cleaned_data.get('DATE_MISSION_DEBUT'),
                    date_mission_fin=form.cleaned_data.get('DATE_MISSION_FIN'),
                    created_by=employe
                )
                messages.success(request, f"Demande d'avance {avance.REFERENCE} créée")
                return redirect('frais:detail_avance', uuid=avance.uuid)
            except Exception as e:
                messages.error(request, str(e))
    else:
        form = AvanceForm()

    return render(request, 'frais/avances/creer.html', {'form': form})


@login_required
def detail_avance(request, uuid):
    """Détail d'une avance."""
    avance = get_object_or_404(NFAV, uuid=uuid)
    employe = request.user.employe

    if avance.EMPLOYE != employe and not _peut_valider(employe):
        return HttpResponseForbidden("Accès non autorisé")

    context = {
        'avance': avance,
        'peut_modifier': avance.peut_etre_modifie() and avance.EMPLOYE == employe,
        'peut_approuver': _peut_valider(employe) and avance.peut_etre_approuve(),
        'peut_verser': _peut_valider(employe) and avance.peut_etre_verse(),
    }

    return render(request, 'frais/avances/detail.html', context)


@login_required
@require_POST
def approuver_avance(request, uuid):
    """Approuve une demande d'avance."""
    avance = get_object_or_404(NFAV, uuid=uuid)
    employe = request.user.employe

    if not _peut_valider(employe):
        return HttpResponseForbidden("Accès non autorisé")

    form = ApprobationAvanceForm(request.POST)
    if form.is_valid():
        try:
            AvanceService.approuver_avance(
                avance,
                employe,
                form.cleaned_data.get('montant_approuve'),
                form.cleaned_data.get('commentaire')
            )
            messages.success(request, f"Avance {avance.REFERENCE} approuvée")
        except Exception as e:
            messages.error(request, str(e))

    return redirect('frais:detail_avance', uuid=uuid)


@login_required
@require_POST
def rejeter_avance(request, uuid):
    """Rejette une demande d'avance."""
    avance = get_object_or_404(NFAV, uuid=uuid)
    employe = request.user.employe

    if not _peut_valider(employe):
        return HttpResponseForbidden("Accès non autorisé")

    form = RejetForm(request.POST)
    if form.is_valid():
        try:
            AvanceService.rejeter_avance(
                avance, employe, form.cleaned_data['commentaire']
            )
            messages.success(request, f"Avance {avance.REFERENCE} rejetée")
        except Exception as e:
            messages.error(request, str(e))
    else:
        messages.error(request, "Un motif de rejet est obligatoire")

    return redirect('frais:detail_avance', uuid=uuid)


@login_required
@require_POST
def verser_avance(request, uuid):
    """Marque une avance comme versée."""
    avance = get_object_or_404(NFAV, uuid=uuid)
    employe = request.user.employe

    if not _peut_valider(employe):
        return HttpResponseForbidden("Accès non autorisé")

    form = VersementAvanceForm(request.POST)
    if form.is_valid():
        try:
            AvanceService.marquer_verse(
                avance,
                form.cleaned_data['date_versement'],
                form.cleaned_data.get('reference_versement')
            )
            messages.success(request, f"Avance {avance.REFERENCE} marquée comme versée")
        except Exception as e:
            messages.error(request, str(e))

    return redirect('frais:detail_avance', uuid=uuid)


# =============================================================================
# VALIDATION (pour valideurs)
# =============================================================================

@login_required
def notes_a_valider(request):
    """Liste des notes en attente de validation."""
    employe = request.user.employe

    if not _peut_valider(employe):
        return HttpResponseForbidden("Accès non autorisé")

    notes = NoteFraisService.get_notes_a_valider()

    paginator = Paginator(notes, 10)
    page = request.GET.get('page', 1)
    notes_page = paginator.get_page(page)

    return render(request, 'frais/validation/notes_a_valider.html', {
        'notes': notes_page
    })


@login_required
def avances_a_approuver(request):
    """Liste des avances en attente d'approbation."""
    employe = request.user.employe

    if not _peut_valider(employe):
        return HttpResponseForbidden("Accès non autorisé")

    avances = AvanceService.get_avances_a_approuver()

    paginator = Paginator(avances, 10)
    page = request.GET.get('page', 1)
    avances_page = paginator.get_page(page)

    return render(request, 'frais/validation/avances_a_approuver.html', {
        'avances': avances_page
    })


@login_required
def notes_a_rembourser(request):
    """Liste des notes validées en attente de remboursement."""
    employe = request.user.employe

    if not _peut_valider(employe):
        return HttpResponseForbidden("Accès non autorisé")

    notes = NoteFraisService.get_notes_a_rembourser()

    paginator = Paginator(notes, 10)
    page = request.GET.get('page', 1)
    notes_page = paginator.get_page(page)

    return render(request, 'frais/validation/notes_a_rembourser.html', {
        'notes': notes_page
    })


# =============================================================================
# STATISTIQUES
# =============================================================================

@login_required
def statistiques_frais(request):
    """Page des statistiques (résultats mis en cache 1 h par employé/année)."""
    employe = request.user.employe
    annee_courante = timezone.now().year
    annee = int(request.GET.get('annee', annee_courante))

    annees_disponibles = list(range(2020, annee_courante + 1))

    # Statistiques personnelles — cache par (employé, année)
    user_cache_key = f'frais_stats_user_{employe.matricule}_{annee}'
    user_stats = cache.get(user_cache_key)
    if user_stats is None:
        user_stats = {
            'mes_stats': StatistiquesFraisService.get_stats_employe(employe, annee),
            'stats_par_categorie': StatistiquesFraisService.get_stats_par_categorie(
                annee, employe
            ),
        }
        cache.set(user_cache_key, user_stats, _CACHE_TTL)

    context = {
        'annee': annee,
        'annees_disponibles': annees_disponibles,
        **user_stats,
    }

    if _peut_valider(employe):
        # Statistiques globales — cache partagé par année (tous les valideurs voient la même chose)
        global_cache_key = f'frais_stats_global_{annee}'
        global_stats = cache.get(global_cache_key)
        if global_stats is None:
            global_stats = {
                'stats_globales': StatistiquesFraisService.get_stats_globales(annee),
                'evolution': StatistiquesFraisService.get_evolution_mensuelle(annee),
                'top_employes': StatistiquesFraisService.get_top_employes(annee),
                'delais': StatistiquesFraisService.get_delai_moyen_traitement(annee),
            }
            cache.set(global_cache_key, global_stats, _CACHE_TTL)
        context.update(global_stats)

    return render(request, 'frais/statistiques.html', context)


# =============================================================================
# UTILITAIRES
# =============================================================================

def _peut_valider(employe) -> bool:
    """Vérifie si un employé peut valider des notes de frais."""
    roles_valideurs = ['DRH', 'RESP_ADMIN', 'DAF', 'ASSISTANT_RH']
    for role in roles_valideurs:
        if employe.has_role(role):
            return True
    return False


@login_required
@require_GET
def api_categories(request):
    """API: Liste des catégories actives."""
    categories = CategorieService.get_categories_actives()
    data = [
        {
            'id': cat.id,
            'code': cat.CODE,
            'libelle': cat.LIBELLE,
            'justificatif_obligatoire': cat.JUSTIFICATIF_OBLIGATOIRE,
            'plafond': str(cat.PLAFOND_DEFAUT) if cat.PLAFOND_DEFAUT else None,
            'icone': cat.ICONE
        }
        for cat in categories
    ]
    return JsonResponse({'categories': data})


# =============================================================================
# GESTION DES CATÉGORIES (Admin/Paramètres)
# =============================================================================

def _peut_gerer_parametres(employe) -> bool:
    """Vérifie si un employé peut gérer les paramètres frais."""
    return employe.has_role('GESTION_APP') or employe.has_role('DRH')


@login_required
def liste_categories(request):
    """Liste des catégories de frais."""
    employe = request.user.employe

    if not _peut_gerer_parametres(employe):
        return HttpResponseForbidden("Accès non autorisé")

    categories = NFCA.objects.all().order_by('ORDRE', 'LIBELLE')

    return render(request, 'frais/categories/liste.html', {
        'categories': categories
    })


@login_required
def creer_categorie(request):
    """Création d'une catégorie de frais."""
    from frais.forms import CategorieForm

    employe = request.user.employe

    if not _peut_gerer_parametres(employe):
        return HttpResponseForbidden("Accès non autorisé")

    if request.method == 'POST':
        form = CategorieForm(request.POST)
        if form.is_valid():
            categorie = form.save()
            messages.success(request, f"Catégorie '{categorie.LIBELLE}' créée avec succès")
            return redirect('frais:liste_categories')
    else:
        form = CategorieForm()

    return render(request, 'frais/categories/form.html', {
        'form': form,
        'titre': 'Nouvelle catégorie de frais'
    })


@login_required
def modifier_categorie(request, pk):
    """Modification d'une catégorie de frais."""
    from frais.forms import CategorieForm

    employe = request.user.employe

    if not _peut_gerer_parametres(employe):
        return HttpResponseForbidden("Accès non autorisé")

    categorie = get_object_or_404(NFCA, pk=pk)

    if request.method == 'POST':
        form = CategorieForm(request.POST, instance=categorie)
        if form.is_valid():
            form.save()
            messages.success(request, f"Catégorie '{categorie.LIBELLE}' modifiée")
            return redirect('frais:liste_categories')
    else:
        form = CategorieForm(instance=categorie)

    return render(request, 'frais/categories/form.html', {
        'form': form,
        'categorie': categorie,
        'titre': f'Modifier la catégorie "{categorie.LIBELLE}"'
    })


@login_required
@require_POST
def supprimer_categorie(request, pk):
    """Suppression d'une catégorie de frais."""
    employe = request.user.employe

    if not _peut_gerer_parametres(employe):
        return HttpResponseForbidden("Accès non autorisé")

    categorie = get_object_or_404(NFCA, pk=pk)

    # Vérifier si la catégorie est utilisée
    if categorie.lignes_frais.exists():
        messages.error(
            request,
            f"Impossible de supprimer '{categorie.LIBELLE}' : elle est utilisée par des notes de frais"
        )
        return redirect('frais:liste_categories')

    libelle = categorie.LIBELLE
    categorie.delete()
    messages.success(request, f"Catégorie '{libelle}' supprimée")
    return redirect('frais:liste_categories')


@login_required
@require_POST
def creer_categories_defaut(request):
    """Crée les catégories par défaut."""
    employe = request.user.employe

    if not _peut_gerer_parametres(employe):
        return HttpResponseForbidden("Accès non autorisé")

    categories_creees = CategorieService.creer_categories_defaut()
    if categories_creees:
        messages.success(request, f"{len(categories_creees)} catégorie(s) par défaut créée(s)")
    else:
        messages.info(request, "Toutes les catégories par défaut existent déjà")

    return redirect('frais:liste_categories')


# =============================================================================
# ADMINISTRATION - NOTES VALIDÉES (DRH, DAF, COMPTABLE)
# =============================================================================

def _peut_administrer_frais(employe) -> bool:
    """Vérifie si un employé peut administrer les frais (DRH, DAF, COMPTABLE)."""
    roles_admin = ['DRH', 'DAF', 'COMPTABLE']
    for role in roles_admin:
        if employe.has_role(role):
            return True
    return False


@login_required
def admin_notes_validees(request):
    """Liste toutes les notes de frais validées avec filtres."""
    employe = request.user.employe

    if not _peut_administrer_frais(employe):
        return HttpResponseForbidden("Accès non autorisé")

    # Récupérer les paramètres de filtre
    mois = request.GET.get('mois')
    annee = request.GET.get('annee', str(timezone.now().year))
    employe_filtre = request.GET.get('employe')
    tri = request.GET.get('tri', '-DATE_VALIDATION')

    # Notes validées et remboursées
    notes = NFNF.objects.filter(
        STATUT__in=['VALIDE', 'REMBOURSE']
    ).select_related('EMPLOYE', 'VALIDEUR')

    # Appliquer les filtres
    if annee:
        notes = notes.filter(DATE_VALIDATION__year=int(annee))

    if mois:
        notes = notes.filter(DATE_VALIDATION__month=int(mois))

    if employe_filtre:
        notes = notes.filter(
            Q(EMPLOYE__matricule__icontains=employe_filtre) |
            Q(EMPLOYE__nom__icontains=employe_filtre) |
            Q(EMPLOYE__prenoms__icontains=employe_filtre)
        )

    # Tri
    notes = notes.order_by(tri)

    # Statistiques
    stats = {
        'total': notes.count(),
        'montant_total': notes.aggregate(Sum('MONTANT_VALIDE'))['MONTANT_VALIDE__sum'] or 0,
        'nb_valide': notes.filter(STATUT='VALIDE').count(),
        'nb_rembourse': notes.filter(STATUT='REMBOURSE').count(),
        'montant_a_rembourser': notes.filter(STATUT='VALIDE').aggregate(
            Sum('MONTANT_VALIDE')
        )['MONTANT_VALIDE__sum'] or 0,
        'montant_rembourse': notes.filter(STATUT='REMBOURSE').aggregate(
            Sum('MONTANT_VALIDE')
        )['MONTANT_VALIDE__sum'] or 0,
    }

    # Pagination
    paginator = Paginator(notes, 20)
    page = request.GET.get('page', 1)
    notes_page = paginator.get_page(page)

    # Liste des années et mois pour les filtres
    annees_disponibles = list(range(2020, timezone.now().year + 1))
    mois_disponibles = [
        (1, 'Janvier'), (2, 'Février'), (3, 'Mars'), (4, 'Avril'),
        (5, 'Mai'), (6, 'Juin'), (7, 'Juillet'), (8, 'Août'),
        (9, 'Septembre'), (10, 'Octobre'), (11, 'Novembre'), (12, 'Décembre')
    ]

    return render(request, 'frais/admin/notes_validees.html', {
        'notes': notes_page,
        'stats': stats,
        'annees_disponibles': annees_disponibles,
        'mois_disponibles': mois_disponibles,
        'annee_selectionnee': annee,
        'mois_selectionne': mois,
        'employe_filtre': employe_filtre,
        'tri': tri,
    })


@login_required
def export_notes_validees_excel(request):
    """Exporte les notes validées en Excel."""
    employe = request.user.employe

    if not _peut_administrer_frais(employe):
        return HttpResponseForbidden("Accès non autorisé")

    # Récupérer les paramètres de filtre
    mois = request.GET.get('mois')
    annee = request.GET.get('annee', str(timezone.now().year))
    employe_filtre = request.GET.get('employe')

    # Notes validées et remboursées
    notes = NFNF.objects.filter(
        STATUT__in=['VALIDE', 'REMBOURSE']
    ).select_related('EMPLOYE', 'VALIDEUR')

    # Appliquer les filtres
    if annee:
        notes = notes.filter(DATE_VALIDATION__year=int(annee))

    if mois:
        notes = notes.filter(DATE_VALIDATION__month=int(mois))

    if employe_filtre:
        notes = notes.filter(
            Q(EMPLOYE__matricule__icontains=employe_filtre) |
            Q(EMPLOYE__nom__icontains=employe_filtre) |
            Q(EMPLOYE__prenoms__icontains=employe_filtre)
        )

    notes = notes.order_by('-DATE_VALIDATION')

    # Créer le workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notes validées"

    # Styles
    header_fill = PatternFill(start_color="1c5d5f", end_color="1c5d5f", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # En-têtes
    headers = [
        'Référence', 'Employé', 'Matricule', 'Période début', 'Période fin',
        'Montant total', 'Montant validé', 'Statut', 'Date validation',
        'Valideur', 'Date remboursement', 'Réf. paiement'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Données
    for row, note in enumerate(notes, 2):
        data = [
            note.REFERENCE,
            f"{note.EMPLOYE.nom} {note.EMPLOYE.prenoms}",
            note.EMPLOYE.matricule,
            note.PERIODE_DEBUT.strftime('%d/%m/%Y') if note.PERIODE_DEBUT else '',
            note.PERIODE_FIN.strftime('%d/%m/%Y') if note.PERIODE_FIN else '',
            float(note.MONTANT_TOTAL),
            float(note.MONTANT_VALIDE),
            note.get_STATUT_display(),
            note.DATE_VALIDATION.strftime('%d/%m/%Y %H:%M') if note.DATE_VALIDATION else '',
            f"{note.VALIDEUR.nom} {note.VALIDEUR.prenoms}" if note.VALIDEUR else '',
            note.DATE_REMBOURSEMENT.strftime('%d/%m/%Y') if note.DATE_REMBOURSEMENT else '',
            note.REFERENCE_PAIEMENT or ''
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border

    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Générer le fichier
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"notes_validees_{annee}"
    if mois:
        filename += f"_{mois:0>2}"
    filename += ".xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


# =============================================================================
# ADMINISTRATION - AVANCES APPROUVÉES (DRH, DAF, COMPTABLE)
# =============================================================================

@login_required
def admin_avances_approuvees(request):
    """Liste toutes les avances approuvées avec filtres."""
    employe = request.user.employe

    if not _peut_administrer_frais(employe):
        return HttpResponseForbidden("Accès non autorisé")

    # Récupérer les paramètres de filtre
    mois = request.GET.get('mois')
    annee = request.GET.get('annee', str(timezone.now().year))
    employe_filtre = request.GET.get('employe')
    statut_filtre = request.GET.get('statut')
    tri = request.GET.get('tri', '-DATE_APPROBATION')

    # Avances approuvées, versées, régularisées
    avances = NFAV.objects.filter(
        STATUT__in=['APPROUVE', 'VERSE', 'REGULARISE']
    ).select_related('EMPLOYE', 'APPROBATEUR')

    # Appliquer les filtres
    if annee:
        avances = avances.filter(DATE_APPROBATION__year=int(annee))

    if mois:
        avances = avances.filter(DATE_APPROBATION__month=int(mois))

    if employe_filtre:
        avances = avances.filter(
            Q(EMPLOYE__matricule__icontains=employe_filtre) |
            Q(EMPLOYE__nom__icontains=employe_filtre) |
            Q(EMPLOYE__prenoms__icontains=employe_filtre)
        )

    if statut_filtre:
        avances = avances.filter(STATUT=statut_filtre)

    # Tri
    avances = avances.order_by(tri)

    # Statistiques
    stats = {
        'total': avances.count(),
        'montant_total': avances.aggregate(Sum('MONTANT_APPROUVE'))['MONTANT_APPROUVE__sum'] or 0,
        'nb_approuve': avances.filter(STATUT='APPROUVE').count(),
        'nb_verse': avances.filter(STATUT='VERSE').count(),
        'nb_regularise': avances.filter(STATUT='REGULARISE').count(),
        'montant_a_verser': avances.filter(STATUT='APPROUVE').aggregate(
            Sum('MONTANT_APPROUVE')
        )['MONTANT_APPROUVE__sum'] or 0,
    }

    # Pagination
    paginator = Paginator(avances, 20)
    page = request.GET.get('page', 1)
    avances_page = paginator.get_page(page)

    # Liste des années et mois pour les filtres
    annees_disponibles = list(range(2020, timezone.now().year + 1))
    mois_disponibles = [
        (1, 'Janvier'), (2, 'Février'), (3, 'Mars'), (4, 'Avril'),
        (5, 'Mai'), (6, 'Juin'), (7, 'Juillet'), (8, 'Août'),
        (9, 'Septembre'), (10, 'Octobre'), (11, 'Novembre'), (12, 'Décembre')
    ]

    return render(request, 'frais/admin/avances_approuvees.html', {
        'avances': avances_page,
        'stats': stats,
        'annees_disponibles': annees_disponibles,
        'mois_disponibles': mois_disponibles,
        'annee_selectionnee': annee,
        'mois_selectionne': mois,
        'employe_filtre': employe_filtre,
        'statut_filtre': statut_filtre,
        'tri': tri,
    })


@login_required
def export_avances_approuvees_excel(request):
    """Exporte les avances approuvées en Excel."""
    employe = request.user.employe

    if not _peut_administrer_frais(employe):
        return HttpResponseForbidden("Accès non autorisé")

    # Récupérer les paramètres de filtre
    mois = request.GET.get('mois')
    annee = request.GET.get('annee', str(timezone.now().year))
    employe_filtre = request.GET.get('employe')
    statut_filtre = request.GET.get('statut')

    # Avances approuvées, versées, régularisées
    avances = NFAV.objects.filter(
        STATUT__in=['APPROUVE', 'VERSE', 'REGULARISE']
    ).select_related('EMPLOYE', 'APPROBATEUR')

    # Appliquer les filtres
    if annee:
        avances = avances.filter(DATE_APPROBATION__year=int(annee))

    if mois:
        avances = avances.filter(DATE_APPROBATION__month=int(mois))

    if employe_filtre:
        avances = avances.filter(
            Q(EMPLOYE__matricule__icontains=employe_filtre) |
            Q(EMPLOYE__nom__icontains=employe_filtre) |
            Q(EMPLOYE__prenoms__icontains=employe_filtre)
        )

    if statut_filtre:
        avances = avances.filter(STATUT=statut_filtre)

    avances = avances.order_by('-DATE_APPROBATION')

    # Créer le workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Avances approuvées"

    # Styles
    header_fill = PatternFill(start_color="1c5d5f", end_color="1c5d5f", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # En-têtes
    headers = [
        'Référence', 'Employé', 'Matricule', 'Montant demandé', 'Montant approuvé',
        'Motif', 'Statut', 'Date approbation', 'Approbateur',
        'Date versement', 'Réf. versement', 'Date régularisation'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Données
    for row, avance in enumerate(avances, 2):
        data = [
            avance.REFERENCE,
            f"{avance.EMPLOYE.nom} {avance.EMPLOYE.prenoms}",
            avance.EMPLOYE.matricule,
            float(avance.MONTANT_DEMANDE),
            float(avance.MONTANT_APPROUVE) if avance.MONTANT_APPROUVE else '',
            avance.MOTIF[:100] if avance.MOTIF else '',
            avance.get_STATUT_display(),
            avance.DATE_APPROBATION.strftime('%d/%m/%Y %H:%M') if avance.DATE_APPROBATION else '',
            f"{avance.APPROBATEUR.nom} {avance.APPROBATEUR.prenoms}" if avance.APPROBATEUR else '',
            avance.DATE_VERSEMENT.strftime('%d/%m/%Y') if avance.DATE_VERSEMENT else '',
            avance.REFERENCE_VERSEMENT or '',
            avance.DATE_REGULARISATION.strftime('%d/%m/%Y') if avance.DATE_REGULARISATION else ''
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border

    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Générer le fichier
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"avances_approuvees_{annee}"
    if mois:
        filename += f"_{mois:0>2}"
    filename += ".xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


# =============================================================================
# VALIDATION REMBOURSEMENT AVEC SIGNATURE
# =============================================================================

@login_required
def admin_remboursements(request):
    """Page de gestion des remboursements avec signature collaborateur."""
    employe = request.user.employe

    if not _peut_administrer_frais(employe):
        return HttpResponseForbidden("Accès non autorisé")

    # Notes validées en attente de remboursement
    notes_a_rembourser = NFNF.objects.filter(
        STATUT='VALIDE'
    ).select_related('EMPLOYE', 'VALIDEUR').order_by('-DATE_VALIDATION')

    # Notes récemment remboursées (30 derniers jours)
    from datetime import timedelta
    date_limite = timezone.now().date() - timedelta(days=30)
    notes_remboursees = NFNF.objects.filter(
        STATUT='REMBOURSE',
        DATE_REMBOURSEMENT__gte=date_limite
    ).select_related('EMPLOYE', 'VALIDEUR').order_by('-DATE_REMBOURSEMENT')[:20]

    # Statistiques
    stats = {
        'nb_a_rembourser': notes_a_rembourser.count(),
        'montant_a_rembourser': notes_a_rembourser.aggregate(
            Sum('MONTANT_VALIDE')
        )['MONTANT_VALIDE__sum'] or 0,
        'nb_rembourse_mois': NFNF.objects.filter(
            STATUT='REMBOURSE',
            DATE_REMBOURSEMENT__year=timezone.now().year,
            DATE_REMBOURSEMENT__month=timezone.now().month
        ).count(),
    }

    return render(request, 'frais/admin/remboursements.html', {
        'notes_a_rembourser': notes_a_rembourser,
        'notes_remboursees': notes_remboursees,
        'stats': stats,
    })


@login_required
@require_POST
def confirmer_remboursement(request, uuid):
    """Confirme le remboursement d'une note avec signature collaborateur."""
    note = get_object_or_404(NFNF, uuid=uuid)
    employe = request.user.employe

    if not _peut_administrer_frais(employe):
        return HttpResponseForbidden("Accès non autorisé")

    if note.STATUT != 'VALIDE':
        messages.error(request, "Cette note ne peut pas être remboursée")
        return redirect('frais:admin_remboursements')

    date_remboursement = request.POST.get('date_remboursement')
    reference_paiement = request.POST.get('reference_paiement')
    signature_confirmee = request.POST.get('signature_confirmee') == 'on'

    if not date_remboursement:
        messages.error(request, "La date de remboursement est obligatoire")
        return redirect('frais:admin_remboursements')

    if not signature_confirmee:
        messages.error(request, "La confirmation de signature est obligatoire")
        return redirect('frais:admin_remboursements')

    try:
        date_obj = datetime.strptime(date_remboursement, '%Y-%m-%d').date()
        NoteFraisService.marquer_rembourse(note, date_obj, reference_paiement)
        messages.success(
            request,
            f"Note {note.REFERENCE} marquée comme remboursée. "
            f"Signature de {note.EMPLOYE.nom} {note.EMPLOYE.prenoms} confirmée."
        )
    except Exception as e:
        messages.error(request, str(e))

    return redirect('frais:admin_remboursements')


@login_required
def fiche_remboursement(request, uuid):
    """Génère une fiche de remboursement à imprimer/signer."""
    note = get_object_or_404(NFNF, uuid=uuid)
    employe = request.user.employe

    # Accès pour l'employé concerné ou les admins
    if note.EMPLOYE != employe and not _peut_administrer_frais(employe):
        return HttpResponseForbidden("Accès non autorisé")

    return render(request, 'frais/admin/fiche_remboursement.html', {
        'note': note,
        'lignes': note.lignes.select_related('CATEGORIE').all(),
    })
