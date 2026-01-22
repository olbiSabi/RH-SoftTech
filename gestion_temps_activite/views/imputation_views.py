# gestion_temps_activite/views/imputation_views.py
"""Vues pour la gestion des imputations de temps (ZDIT)."""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum
from django.core.paginator import Paginator
from django.utils import timezone
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

from absence.decorators import role_required
from employee.models import ZY00
from gestion_temps_activite.models import ZDPJ, ZDTA, ZDAC, ZDIT
from gestion_temps_activite.forms import ZDITForm, RechercheImputationForm
from gestion_temps_activite.services import ImputationService, StatistiqueService


@login_required
def imputation_liste(request):
    """Liste des imputations de temps."""
    imputations = ZDIT.objects.select_related(
        'employe', 'tache', 'tache__projet', 'activite'
    ).all().order_by('-date', '-date_creation')

    # Formulaire de recherche
    form = RechercheImputationForm(request.GET or None)

    if form.is_valid():
        data = form.cleaned_data
        if data.get('employe'):
            imputations = imputations.filter(employe=data['employe'])
        if data.get('projet'):
            imputations = imputations.filter(tache__projet=data['projet'])
        if data.get('tache'):
            imputations = imputations.filter(tache=data['tache'])
        if data.get('activite'):
            imputations = imputations.filter(activite=data['activite'])
        if data.get('date_debut'):
            imputations = imputations.filter(date__gte=data['date_debut'])
        if data.get('date_fin'):
            imputations = imputations.filter(date__lte=data['date_fin'])
        if data.get('valide'):
            imputations = imputations.filter(valide=(data['valide'] == 'True'))
        if data.get('facture'):
            imputations = imputations.filter(facture=(data['facture'] == 'True'))

    # Statistiques agrégées
    stats = StatistiqueService.get_stats_imputations(imputations)

    # Pagination
    paginator = Paginator(imputations, 25)
    page_number = request.GET.get('page')
    imputations_page = paginator.get_page(page_number)

    context = {
        'imputations': imputations_page,
        'form': form,
        'total_heures': stats['total_heures'],
        'heures_validees': stats['heures_validees'],
        'heures_facturables': stats['heures_facturables']
    }

    return render(request, 'gestion_temps_activite/imputation_liste.html', context)


@login_required
def imputation_mes_temps(request):
    """Liste des imputations de l'utilisateur connecté."""
    if not hasattr(request.user, 'employe'):
        messages.error(request, "Vous n'avez pas de profil employé associé.")
        return redirect('gestion_temps_activite:dashboard')

    employe = request.user.employe
    imputations = ZDIT.objects.filter(employe=employe).select_related(
        'tache', 'tache__projet', 'activite'
    ).order_by('-date', '-date_creation')

    # Filtres basiques
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')

    if date_debut:
        imputations = imputations.filter(date__gte=date_debut)
    if date_fin:
        imputations = imputations.filter(date__lte=date_fin)

    # Heures du mois courant
    heures_mois = ImputationService.get_heures_mois_courant(employe)

    # Pagination
    paginator = Paginator(imputations, 25)
    page_number = request.GET.get('page')
    imputations_page = paginator.get_page(page_number)

    context = {
        'imputations': imputations_page,
        'heures_mois': heures_mois,
        'date_debut': date_debut,
        'date_fin': date_fin
    }

    return render(request, 'gestion_temps_activite/imputation_mes_temps.html', context)


@login_required
def imputation_create(request):
    """Créer une nouvelle imputation de temps."""
    if request.method == 'POST':
        form = ZDITForm(request.POST, user=request.user)
        if form.is_valid():
            imputation = form.save(commit=False)
            if hasattr(request.user, 'employe'):
                imputation.employe = request.user.employe
            imputation.save()

            messages.success(request, 'Imputation enregistrée avec succès.')
            return redirect('gestion_temps_activite:imputation_mes_temps')
        else:
            messages.error(request, 'Erreur lors de l\'enregistrement.')
    else:
        tache_id = request.GET.get('tache')
        initial = {'date': timezone.now().date()}
        if tache_id:
            initial['tache'] = tache_id
        form = ZDITForm(initial=initial, user=request.user)

    context = {
        'form': form,
        'title': 'Nouvelle Imputation',
        'action': 'Créer'
    }

    return render(request, 'gestion_temps_activite/imputation_form.html', context)


@login_required
def imputation_update(request, pk):
    """Modifier une imputation."""
    imputation = get_object_or_404(ZDIT, pk=pk)

    # Vérifier les permissions
    if hasattr(request.user, 'employe'):
        if not ImputationService.peut_modifier(imputation, request.user.employe):
            messages.error(request, "Vous ne pouvez pas modifier cette imputation.")
            return redirect('gestion_temps_activite:imputation_mes_temps')

    if request.method == 'POST':
        form = ZDITForm(request.POST, instance=imputation, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Imputation modifiée avec succès.')
            return redirect('gestion_temps_activite:imputation_mes_temps')
        else:
            messages.error(request, 'Erreur lors de la modification.')
    else:
        form = ZDITForm(instance=imputation, user=request.user)

    context = {
        'form': form,
        'imputation': imputation,
        'title': 'Modifier Imputation',
        'action': 'Modifier'
    }

    return render(request, 'gestion_temps_activite/imputation_form.html', context)


@login_required
def imputation_delete(request, pk):
    """Supprimer une imputation."""
    imputation = get_object_or_404(ZDIT, pk=pk)

    # Vérifier les permissions
    if hasattr(request.user, 'employe'):
        if not ImputationService.peut_supprimer(imputation, request.user.employe):
            messages.error(request, "Vous ne pouvez pas supprimer cette imputation.")
            return redirect('gestion_temps_activite:imputation_mes_temps')

    if request.method == 'POST':
        try:
            imputation.delete()
            messages.success(request, 'Imputation supprimée avec succès.')
        except Exception as e:
            messages.error(request, f'Erreur lors de la suppression : {str(e)}')
        return redirect('gestion_temps_activite:imputation_mes_temps')

    context = {
        'imputation': imputation
    }

    return render(request, 'gestion_temps_activite/imputation_confirm_delete.html', context)


@role_required('MANAGER', 'DRH', 'GESTION_APP', 'DIRECTEUR')
@login_required
def imputation_validation(request):
    """Liste des imputations à valider."""
    imputations = ZDIT.objects.filter(valide=False).select_related(
        'employe', 'tache', 'tache__projet', 'activite'
    ).order_by('-date')

    # Filtres
    employe_filter = request.GET.get('employe')
    if employe_filter:
        imputations = imputations.filter(employe_id=employe_filter)

    projet_filter = request.GET.get('projet')
    if projet_filter:
        imputations = imputations.filter(tache__projet_id=projet_filter)

    # Pagination
    paginator = Paginator(imputations, 25)
    page_number = request.GET.get('page')
    imputations_page = paginator.get_page(page_number)

    # Listes pour filtres
    employes = ZY00.objects.filter(etat='actif').order_by('nom', 'prenoms')
    projets = ZDPJ.objects.filter(actif=True).order_by('nom_projet')

    context = {
        'imputations': imputations_page,
        'employes': employes,
        'projets': projets,
        'employe_filter': employe_filter,
        'projet_filter': projet_filter
    }

    return render(request, 'gestion_temps_activite/imputation_validation.html', context)


@role_required('MANAGER', 'DRH', 'GESTION_APP', 'DIRECTEUR')
@login_required
def imputation_valider(request, pk):
    """Valider une imputation."""
    imputation = get_object_or_404(ZDIT, pk=pk)

    if hasattr(request.user, 'employe'):
        if ImputationService.valider_imputation(imputation, request.user.employe):
            messages.success(request, 'Imputation validée avec succès.')
        else:
            messages.error(request, 'Cette imputation est déjà validée.')

    return redirect('gestion_temps_activite:imputation_validation')


@role_required('MANAGER', 'DRH', 'GESTION_APP', 'DIRECTEUR')
@login_required
def imputation_rejeter(request, pk):
    """Rejeter une imputation."""
    imputation = get_object_or_404(ZDIT, pk=pk)

    if request.method == 'POST':
        motif = request.POST.get('motif', '')
        if ImputationService.rejeter_imputation(imputation, motif):
            messages.success(request, 'Imputation rejetée.')
        else:
            messages.error(request, 'Veuillez fournir un motif de rejet.')

    return redirect('gestion_temps_activite:imputation_validation')


@role_required('MANAGER', 'DRH', 'GESTION_APP', 'DIRECTEUR')
@login_required
def imputation_export_excel(request):
    """Exporter les imputations en Excel."""
    # Récupérer les filtres
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    projet_id = request.GET.get('projet')

    imputations = ZDIT.objects.select_related(
        'employe', 'tache', 'tache__projet', 'activite'
    ).filter(valide=True).order_by('date')

    if date_debut:
        imputations = imputations.filter(date__gte=date_debut)
    if date_fin:
        imputations = imputations.filter(date__lte=date_fin)
    if projet_id:
        imputations = imputations.filter(tache__projet_id=projet_id)

    # Créer le workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Imputations"

    # En-têtes
    headers = [
        'Date', 'Employé', 'Projet', 'Tâche', 'Activité',
        'Durée (h)', 'Facturable', 'Taux Horaire', 'Montant', 'Commentaire'
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Données
    for row, imp in enumerate(imputations, 2):
        montant = ImputationService.calculer_montant_facturable(imp)
        ws.cell(row=row, column=1, value=imp.date.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=f"{imp.employe.nom} {imp.employe.prenoms}" if imp.employe else '')
        ws.cell(row=row, column=3, value=imp.tache.projet.nom_projet if imp.tache and imp.tache.projet else '')
        ws.cell(row=row, column=4, value=imp.tache.titre if imp.tache else '')
        ws.cell(row=row, column=5, value=imp.activite.libelle if imp.activite else '')
        ws.cell(row=row, column=6, value=float(imp.duree) if imp.duree else 0)
        ws.cell(row=row, column=7, value='Oui' if imp.facturable else 'Non')
        ws.cell(row=row, column=8, value=float(imp.taux_horaire_applique) if imp.taux_horaire_applique else 0)
        ws.cell(row=row, column=9, value=montant)
        ws.cell(row=row, column=10, value=imp.commentaire or '')

    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"imputations_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response
