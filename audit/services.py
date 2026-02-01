# audit/services.py
"""
Services pour le module Conformité & Audit.
"""
import io
from decimal import Decimal
from datetime import date, timedelta
from django.db.models import Count, Q, F
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings

from .models import AUAL, AURA, AURC
from core.models import ZDLOG


class ConformiteService:
    """Service de vérification de conformité."""

    @staticmethod
    def verifier_contrats_expirants(jours_avant=None):
        """
        Vérifie les contrats qui expirent bientôt selon les règles de conformité.
        Génère des alertes pour les contrats proches de l'expiration.

        Args:
            jours_avant: Nombre de jours avant expiration (optionnel).
                        Si None, utilise les règles AURC de type CONTRAT.

        Returns:
            Liste des alertes créées.
        """
        from employee.models import ZYCO

        alertes_creees = []

        # Récupérer les règles de conformité actives pour les contrats
        regles_contrat = AURC.objects.filter(
            TYPE_REGLE='CONTRAT',
            STATUT=True
        )

        # Si aucune règle n'existe, utiliser un comportement par défaut
        if not regles_contrat.exists():
            if jours_avant is None:
                jours_avant = 30  # Valeur par défaut
            regles_contrat = [None]  # Liste vide pour itérer au moins une fois

        for regle in regles_contrat:
            # Déterminer le nombre de jours avant expiration
            if regle:
                jours_avant_expiration = regle.JOURS_AVANT_EXPIRATION
                severite = regle.SEVERITE
            else:
                jours_avant_expiration = jours_avant
                severite = 'WARNING'

            date_limite = timezone.now().date() + timedelta(days=jours_avant_expiration)

            contrats_expirants = ZYCO.objects.filter(
                actif=True,
                date_fin__isnull=False,
                date_fin__lte=date_limite,
                date_fin__gte=timezone.now().date()
            ).select_related('employe')

            for contrat in contrats_expirants:
                # Vérifier si une alerte existe déjà pour ce contrat et cette règle
                alerte_existante = AUAL.objects.filter(
                    TYPE_ALERTE='CONTRAT',
                    TABLE_REFERENCE='ZYCO',
                    RECORD_ID=str(contrat.pk),
                    STATUT__in=['NOUVEAU', 'EN_COURS']
                )

                # Si une règle existe, vérifier aussi qu'elle correspond à cette règle
                if regle:
                    alerte_existante = alerte_existante.filter(REGLE=regle)

                if not alerte_existante.exists():
                    jours_restants = (contrat.date_fin - timezone.now().date()).days

                    # Déterminer la priorité en fonction des jours restants
                    if jours_restants <= 7:
                        priorite = 'CRITIQUE'
                    elif jours_restants <= 14:
                        priorite = 'HAUTE'
                    elif jours_restants <= 30:
                        priorite = 'MOYENNE'
                    else:
                        priorite = 'BASSE'

                    # Créer l'alerte
                    alerte = AUAL.objects.create(
                        REGLE=regle,
                        TYPE_ALERTE='CONTRAT',
                        TITRE=f"Contrat expirant - {contrat.employe.nom} {contrat.employe.prenoms}",
                        DESCRIPTION=f"Le contrat {contrat.type_contrat} de {contrat.employe.nom} {contrat.employe.prenoms} "
                                   f"expire le {contrat.date_fin.strftime('%d/%m/%Y')} ({jours_restants} jours restants).\n\n"
                                   f"Type de contrat: {contrat.type_contrat}\n"
                                   f"Date de début: {contrat.date_debut.strftime('%d/%m/%Y')}\n"
                                   f"Date de fin: {contrat.date_fin.strftime('%d/%m/%Y')}",
                        PRIORITE=priorite,
                        EMPLOYE=contrat.employe,
                        TABLE_REFERENCE='ZYCO',
                        RECORD_ID=str(contrat.pk),
                        DATE_ECHEANCE=contrat.date_fin
                    )
                    alertes_creees.append(alerte)

                    # Envoyer les notifications si configurées dans la règle
                    if regle:
                        ConformiteService._envoyer_notifications_alerte(alerte, regle)

        return alertes_creees

    @staticmethod
    def _envoyer_notifications_alerte(alerte, regle):
        """
        Envoie les notifications configurées pour une alerte.

        Args:
            alerte: L'alerte créée
            regle: La règle de conformité associée
        """
        destinataires = []

        # Notifier l'employé
        if regle.NOTIFIER_EMPLOYE and alerte.EMPLOYE and hasattr(alerte.EMPLOYE, 'user'):
            try:
                if alerte.EMPLOYE.user.email:
                    destinataires.append(alerte.EMPLOYE.user.email)
            except:
                pass

        # Notifier le manager
        if regle.NOTIFIER_MANAGER and alerte.EMPLOYE:
            try:
                # Récupérer le manager de l'employé
                from employee.services.hierarchy_service import HierarchyService
                manager = HierarchyService.get_manager_direct(alerte.EMPLOYE)
                if manager and hasattr(manager, 'user') and manager.user.email:
                    destinataires.append(manager.user.email)
            except:
                pass

        # Notifier les RH
        if regle.NOTIFIER_RH:
            try:
                from employee.models import ZY00
                # Récupérer les employés avec le rôle DRH ou ASSISTANT_RH
                rh_users = ZY00.objects.filter(
                    Q(roles__role__code='DRH') | Q(roles__role__code='ASSISTANT_RH'),
                    etat='actif'
                ).distinct()

                for rh in rh_users:
                    if hasattr(rh, 'user') and rh.user.email:
                        destinataires.append(rh.user.email)
            except:
                pass

        # Ajouter les emails supplémentaires
        if regle.EMAILS_SUPPLEMENTAIRES:
            emails_supp = [email.strip() for email in regle.EMAILS_SUPPLEMENTAIRES.split('\n') if email.strip()]
            destinataires.extend(emails_supp)

        # Envoyer l'email si il y a des destinataires
        if destinataires:
            destinataires = list(set(destinataires))  # Supprimer les doublons

            sujet = f"[{alerte.PRIORITE}] {alerte.TITRE}"
            message = f"""
Bonjour,

Une nouvelle alerte de conformité a été générée :

Type: {alerte.TYPE_ALERTE}
Priorité: {alerte.PRIORITE}
Employé concerné: {alerte.EMPLOYE.nom} {alerte.EMPLOYE.prenoms} ({alerte.EMPLOYE.matricule})

{alerte.DESCRIPTION}

Veuillez prendre les mesures nécessaires dans les meilleurs délais.

Cordialement,
Le système ONIAN-EasyM
            """

            try:
                send_mail(
                    subject=sujet,
                    message=message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=destinataires,
                    fail_silently=True,
                )

                # Marquer la notification comme envoyée
                alerte.NOTIFICATION_ENVOYEE = True
                alerte.DATE_NOTIFICATION = timezone.now()
                alerte.save(update_fields=['NOTIFICATION_ENVOYEE', 'DATE_NOTIFICATION'])
            except Exception as e:
                # Logger l'erreur mais ne pas bloquer la création de l'alerte
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Erreur lors de l'envoi de notification pour l'alerte {alerte.REFERENCE}: {str(e)}")

    @staticmethod
    def verifier_documents_manquants():
        """
        Vérifie les documents obligatoires manquants.
        """
        from employee.models import ZY00, ZYDO

        # Liste des documents obligatoires par défaut
        documents_obligatoires = [
            ('CNI', 'Carte Nationale d\'Identité'),
            ('CV', 'Curriculum Vitae'),
            ('DIPLOME', 'Diplôme'),
            ('RIB', 'Relevé d\'Identité Bancaire'),
        ]

        alertes_creees = []
        employes_actifs = ZY00.objects.filter(etat='actif')

        for employe in employes_actifs:
            for code_doc, libelle_doc in documents_obligatoires:
                # Vérifier si le document existe
                doc_existe = ZYDO.objects.filter(
                    MATRICULE=employe,
                    TYPE_DOCUMENT=code_doc,
                    STATUT=True
                ).exists()

                if not doc_existe:
                    # Vérifier si une alerte existe déjà
                    alerte_existante = AUAL.objects.filter(
                        TYPE_ALERTE='DOCUMENT',
                        EMPLOYE=employe,
                        TITRE__icontains=code_doc,
                        STATUT__in=['NOUVEAU', 'EN_COURS']
                    ).exists()

                    if not alerte_existante:
                        alerte = AUAL.objects.create(
                            TYPE_ALERTE='DOCUMENT',
                            TITRE=f"Document manquant - {libelle_doc}",
                            DESCRIPTION=f"Le document '{libelle_doc}' est manquant pour "
                                       f"{employe.nom} {employe.prenoms} ({employe.matricule}).",
                            PRIORITE='MOYENNE',
                            EMPLOYE=employe
                        )
                        alertes_creees.append(alerte)

        return alertes_creees

    @staticmethod
    def verifier_visites_medicales():
        """
        Vérifie les visites médicales expirées ou à venir.
        """
        from employee.models import ZY00

        alertes_creees = []
        date_limite = timezone.now().date() + timedelta(days=30)

        employes = ZY00.objects.filter(
            etat='actif',
            date_visite_medicale__isnull=False,
            date_visite_medicale__lte=date_limite
        )

        for employe in employes:
            alerte_existante = AUAL.objects.filter(
                TYPE_ALERTE='VISITE_MEDICALE',
                EMPLOYE=employe,
                STATUT__in=['NOUVEAU', 'EN_COURS']
            ).exists()

            if not alerte_existante:
                jours = (employe.date_visite_medicale - timezone.now().date()).days
                if jours < 0:
                    priorite = 'CRITIQUE'
                    titre = f"Visite médicale expirée - {employe.nom} {employe.prenoms}"
                else:
                    priorite = 'HAUTE' if jours <= 7 else 'MOYENNE'
                    titre = f"Visite médicale à renouveler - {employe.nom} {employe.prenoms}"

                alerte = AUAL.objects.create(
                    TYPE_ALERTE='VISITE_MEDICALE',
                    TITRE=titre,
                    DESCRIPTION=f"La visite médicale de {employe.nom} {employe.prenoms} "
                               f"{'a expiré' if jours < 0 else 'expire'} le {employe.date_visite_medicale.strftime('%d/%m/%Y')}.",
                    PRIORITE=priorite,
                    EMPLOYE=employe,
                    DATE_ECHEANCE=employe.date_visite_medicale
                )
                alertes_creees.append(alerte)

        return alertes_creees

    @staticmethod
    def verifier_materiel_en_retard():
        """
        Vérifie les prêts de matériel en retard.
        """
        from materiel.models import MTAF

        alertes_creees = []
        today = timezone.now().date()

        affectations_retard = MTAF.objects.filter(
            TYPE_AFFECTATION='PRET',
            ACTIF=True,
            DATE_FIN__isnull=True,
            DATE_RETOUR_PREVUE__lt=today
        ).select_related('MATERIEL', 'EMPLOYE')

        for affectation in affectations_retard:
            alerte_existante = AUAL.objects.filter(
                TYPE_ALERTE='MATERIEL',
                TABLE_REFERENCE='MTAF',
                RECORD_ID=str(affectation.pk),
                STATUT__in=['NOUVEAU', 'EN_COURS']
            ).exists()

            if not alerte_existante:
                jours_retard = (today - affectation.DATE_RETOUR_PREVUE).days
                alerte = AUAL.objects.create(
                    TYPE_ALERTE='MATERIEL',
                    TITRE=f"Prêt matériel en retard - {affectation.MATERIEL.CODE_INTERNE}",
                    DESCRIPTION=f"Le matériel {affectation.MATERIEL.DESIGNATION} prêté à "
                               f"{affectation.EMPLOYE.nom} {affectation.EMPLOYE.prenoms} devait être retourné "
                               f"le {affectation.DATE_RETOUR_PREVUE.strftime('%d/%m/%Y')} ({jours_retard} jours de retard).",
                    PRIORITE='HAUTE' if jours_retard > 7 else 'MOYENNE',
                    EMPLOYE=affectation.EMPLOYE,
                    TABLE_REFERENCE='MTAF',
                    RECORD_ID=str(affectation.pk),
                    DATE_ECHEANCE=affectation.DATE_RETOUR_PREVUE
                )
                alertes_creees.append(alerte)

        return alertes_creees

    @staticmethod
    def executer_toutes_verifications():
        """
        Exécute toutes les vérifications de conformité.
        """
        resultats = {
            'contrats': [],
            'documents': [],
            'visites_medicales': [],
            'materiel': [],
        }

        try:
            resultats['contrats'] = ConformiteService.verifier_contrats_expirants()
        except Exception as e:
            resultats['contrats_erreur'] = str(e)

        try:
            resultats['documents'] = ConformiteService.verifier_documents_manquants()
        except Exception as e:
            resultats['documents_erreur'] = str(e)

        try:
            resultats['visites_medicales'] = ConformiteService.verifier_visites_medicales()
        except Exception as e:
            resultats['visites_medicales_erreur'] = str(e)

        try:
            resultats['materiel'] = ConformiteService.verifier_materiel_en_retard()
        except Exception as e:
            resultats['materiel_erreur'] = str(e)

        return resultats


class AlerteService:
    """Service de gestion des alertes."""

    @staticmethod
    def get_alertes_dashboard():
        """Retourne les statistiques des alertes pour le dashboard."""
        stats = AUAL.objects.aggregate(
            total=Count('id'),
            nouveaux=Count('id', filter=Q(STATUT='NOUVEAU')),
            en_cours=Count('id', filter=Q(STATUT='EN_COURS')),
            resolus=Count('id', filter=Q(STATUT='RESOLU')),
            critiques=Count('id', filter=Q(PRIORITE='CRITIQUE', STATUT__in=['NOUVEAU', 'EN_COURS'])),
            en_retard=Count('id', filter=Q(
                DATE_ECHEANCE__lt=timezone.now().date(),
                STATUT__in=['NOUVEAU', 'EN_COURS']
            ))
        )
        return stats

    @staticmethod
    def get_alertes_par_type():
        """Retourne le nombre d'alertes actives par type avec libellés."""
        from audit.models import AURC

        # Récupérer les stats par type
        stats = AUAL.objects.filter(
            STATUT__in=['NOUVEAU', 'EN_COURS']
        ).values('TYPE_ALERTE').annotate(
            count=Count('id')
        ).order_by('-count')

        # Créer un dictionnaire de mapping code -> libellé
        type_dict = dict(AURC.TYPE_CHOICES)

        # Transformer les codes en libellés
        result = []
        for item in stats:
            code = item['TYPE_ALERTE']
            result.append({
                'TYPE_ALERTE': type_dict.get(code, code),  # Utilise le libellé si disponible
                'count': item['count']
            })

        return result

    @staticmethod
    def resoudre_alerte(alerte, employe, commentaire=''):
        """Résout une alerte."""
        alerte.STATUT = 'RESOLU'
        alerte.DATE_RESOLUTION = timezone.now()
        alerte.RESOLU_PAR = employe
        alerte.COMMENTAIRE_RESOLUTION = commentaire
        alerte.save()
        return alerte

    @staticmethod
    def assigner_alerte(alerte, employe_assigne):
        """Assigne une alerte à un employé."""
        alerte.ASSIGNE_A = employe_assigne
        alerte.STATUT = 'EN_COURS'
        alerte.save()
        return alerte

    @staticmethod
    def ignorer_alerte(alerte, employe, commentaire=''):
        """Ignore une alerte."""
        alerte.STATUT = 'IGNORE'
        alerte.DATE_RESOLUTION = timezone.now()
        alerte.RESOLU_PAR = employe
        alerte.COMMENTAIRE_RESOLUTION = commentaire
        alerte.save()
        return alerte


class LogService:
    """Service de consultation des logs."""

    @staticmethod
    def get_logs_recents(limit=100):
        """Retourne les logs les plus récents."""
        return ZDLOG.objects.all()[:limit]

    @staticmethod
    def get_logs_par_table(table_name, limit=100):
        """Retourne les logs d'une table spécifique."""
        return ZDLOG.objects.filter(TABLE_NAME=table_name)[:limit]

    @staticmethod
    def get_logs_par_utilisateur(user, limit=100):
        """Retourne les logs d'un utilisateur."""
        return ZDLOG.objects.filter(USER=user)[:limit]

    @staticmethod
    def get_logs_par_periode(date_debut, date_fin):
        """Retourne les logs d'une période."""
        return ZDLOG.objects.filter(
            DATE_MODIFICATION__date__gte=date_debut,
            DATE_MODIFICATION__date__lte=date_fin
        )

    @staticmethod
    def get_statistiques_logs(date_debut=None, date_fin=None):
        """Retourne des statistiques sur les logs."""
        queryset = ZDLOG.objects.all()

        if date_debut:
            queryset = queryset.filter(DATE_MODIFICATION__date__gte=date_debut)
        if date_fin:
            queryset = queryset.filter(DATE_MODIFICATION__date__lte=date_fin)

        stats = queryset.aggregate(
            total=Count('id'),
            creations=Count('id', filter=Q(TYPE_MOUVEMENT='CREATE')),
            modifications=Count('id', filter=Q(TYPE_MOUVEMENT='UPDATE')),
            suppressions=Count('id', filter=Q(TYPE_MOUVEMENT='DELETE')),
        )

        # Logs par table
        stats['par_table'] = queryset.values('TABLE_NAME').annotate(
            count=Count('id')
        ).order_by('-count')[:10]

        # Logs par utilisateur
        stats['par_utilisateur'] = queryset.exclude(
            USER_NAME=''
        ).values('USER_NAME').annotate(
            count=Count('id')
        ).order_by('-count')[:10]

        return stats


class RapportAuditService:
    """Service de génération de rapports d'audit."""

    @staticmethod
    def generer_rapport_conformite(date_debut, date_fin, genere_par, format_export='PDF'):
        """Génère un rapport de conformité."""
        rapport = AURA.objects.create(
            TITRE=f"Rapport de conformité du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}",
            TYPE_RAPPORT='CONFORMITE',
            FORMAT=format_export,
            DATE_DEBUT=date_debut,
            DATE_FIN=date_fin,
            GENERE_PAR=genere_par
        )

        try:
            # Collecter les données
            alertes = AUAL.objects.filter(
                DATE_DETECTION__date__gte=date_debut,
                DATE_DETECTION__date__lte=date_fin
            )

            resume = {
                'total_alertes': alertes.count(),
                'par_type': list(alertes.values('TYPE_ALERTE').annotate(count=Count('id'))),
                'par_priorite': list(alertes.values('PRIORITE').annotate(count=Count('id'))),
                'par_statut': list(alertes.values('STATUT').annotate(count=Count('id'))),
                'resolues': alertes.filter(STATUT='RESOLU').count(),
                'en_cours': alertes.filter(STATUT='EN_COURS').count(),
            }

            rapport.NB_ENREGISTREMENTS = alertes.count()
            rapport.RESUME = resume
            rapport.STATUT = 'TERMINE'
            rapport.save()

        except Exception as e:
            rapport.STATUT = 'ERREUR'
            rapport.MESSAGE_ERREUR = str(e)
            rapport.save()

        return rapport

    @staticmethod
    def generer_rapport_logs(date_debut, date_fin, genere_par, format_export='EXCEL', filtres=None):
        """Génère un rapport des logs d'activité."""
        rapport = AURA.objects.create(
            TITRE=f"Rapport des logs du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}",
            TYPE_RAPPORT='LOGS',
            FORMAT=format_export,
            DATE_DEBUT=date_debut,
            DATE_FIN=date_fin,
            GENERE_PAR=genere_par,
            FILTRES=filtres
        )

        try:
            logs = ZDLOG.objects.filter(
                DATE_MODIFICATION__date__gte=date_debut,
                DATE_MODIFICATION__date__lte=date_fin
            )

            if filtres:
                if filtres.get('table_name'):
                    logs = logs.filter(TABLE_NAME=filtres['table_name'])
                if filtres.get('type_mouvement'):
                    logs = logs.filter(TYPE_MOUVEMENT=filtres['type_mouvement'])
                if filtres.get('user_id'):
                    logs = logs.filter(USER_id=filtres['user_id'])

            resume = {
                'total_logs': logs.count(),
                'par_type': list(logs.values('TYPE_MOUVEMENT').annotate(count=Count('id'))),
                'par_table': list(logs.values('TABLE_NAME').annotate(count=Count('id')).order_by('-count')[:10]),
            }

            rapport.NB_ENREGISTREMENTS = logs.count()
            rapport.RESUME = resume
            rapport.STATUT = 'TERMINE'
            rapport.save()

        except Exception as e:
            rapport.STATUT = 'ERREUR'
            rapport.MESSAGE_ERREUR = str(e)
            rapport.save()

        return rapport

    @staticmethod
    def generer_rapport_contrats(date_debut, date_fin, genere_par, format_export='EXCEL'):
        """Génère un rapport sur les contrats."""
        from employee.models import ZYCO

        rapport = AURA.objects.create(
            TITRE=f"Rapport des contrats du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}",
            TYPE_RAPPORT='CONTRATS',
            FORMAT=format_export,
            DATE_DEBUT=date_debut,
            DATE_FIN=date_fin,
            GENERE_PAR=genere_par
        )

        try:
            # Contrats actifs
            contrats_actifs = ZYCO.objects.filter(ACTIF=True)

            # Contrats expirant dans la période
            contrats_expirants = contrats_actifs.filter(
                DATE_FIN__gte=date_debut,
                DATE_FIN__lte=date_fin
            )

            # Contrats créés dans la période
            contrats_crees = ZYCO.objects.filter(
                DATE_DEBUT__gte=date_debut,
                DATE_DEBUT__lte=date_fin
            )

            resume = {
                'total_contrats_actifs': contrats_actifs.count(),
                'contrats_expirants': contrats_expirants.count(),
                'contrats_crees': contrats_crees.count(),
                'par_type': list(contrats_actifs.values('TYPE_CONTRAT').annotate(count=Count('id'))),
            }

            rapport.NB_ENREGISTREMENTS = contrats_actifs.count()
            rapport.RESUME = resume
            rapport.STATUT = 'TERMINE'
            rapport.save()

        except Exception as e:
            rapport.STATUT = 'ERREUR'
            rapport.MESSAGE_ERREUR = str(e)
            rapport.save()

        return rapport

    @staticmethod
    def exporter_rapport_excel(rapport):
        """Exporte un rapport au format Excel."""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1c5d5f", end_color="1c5d5f", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # En-tête du rapport
        ws.merge_cells('A1:D1')
        ws['A1'] = rapport.TITRE
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")

        ws['A2'] = f"Généré le: {rapport.DATE_GENERATION.strftime('%d/%m/%Y %H:%M')}"
        ws['A3'] = f"Période: {rapport.DATE_DEBUT.strftime('%d/%m/%Y')} au {rapport.DATE_FIN.strftime('%d/%m/%Y')}"

        # Résumé
        if rapport.RESUME:
            row = 5
            ws[f'A{row}'] = "RÉSUMÉ"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1

            for key, value in rapport.RESUME.items():
                if not isinstance(value, list):
                    ws[f'A{row}'] = key.replace('_', ' ').title()
                    ws[f'B{row}'] = str(value)
                    row += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
