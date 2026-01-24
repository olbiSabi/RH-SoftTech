# materiel/views.py
"""
Vues pour le module Suivi du Matériel & Parc.
"""
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils import timezone
from django.views.decorators.http import require_POST, require_GET
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from materiel.models import MTCA, MTFO, MTMT, MTAF, MTMV, MTMA
from materiel.forms import (
    MTCAForm, MTFOForm, MTMTForm, MTMTEditForm,
    AffectationForm, RetourForm, MTMAForm, TerminerMaintenanceForm,
    ReformeForm, FiltresMaterielForm, FiltresMaintenanceForm
)
from materiel.services import MaterielService, StatistiquesMaterielService
from employee.models import ZY00


def _peut_gerer_materiel(employe):
    """Vérifie si l'employé peut gérer le matériel."""
    if not employe:
        return False
    return employe.has_role('DRH') or employe.has_role('GESTION_APP') or employe.has_role('RESP_ADMIN')


def _peut_affecter_materiel(employe):
    """Vérifie si l'employé peut affecter du matériel."""
    if not employe:
        return False
    return _peut_gerer_materiel(employe) or employe.has_role('ASSISTANT_RH')


# ============================================================================
# DASHBOARD
# ============================================================================

@login_required
def dashboard(request):
    """Dashboard du module matériel."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_gerer_materiel(employe) and not _peut_affecter_materiel(employe):
        messages.error(request, "Vous n'avez pas accès à ce module.")
        return redirect('home')

    stats = StatistiquesMaterielService.get_stats_globales()
    stats_categories = StatistiquesMaterielService.get_stats_par_categorie()
    alertes = StatistiquesMaterielService.get_alertes()
    valeur_parc = StatistiquesMaterielService.get_valeur_parc()

    # Dernières acquisitions
    dernieres_acquisitions = MTMT.objects.exclude(
        STATUT='REFORME'
    ).order_by('-DATE_ACQUISITION')[:5]

    # Maintenances à venir
    maintenances_a_venir = MTMA.objects.filter(
        STATUT='PLANIFIE',
        DATE_PLANIFIEE__gte=timezone.now().date()
    ).order_by('DATE_PLANIFIEE')[:5]

    context = {
        'stats': stats,
        'stats_categories': stats_categories,
        'alertes': alertes,
        'valeur_parc': valeur_parc,
        'dernieres_acquisitions': dernieres_acquisitions,
        'maintenances_a_venir': maintenances_a_venir,
        'peut_gerer': _peut_gerer_materiel(employe),
    }
    return render(request, 'materiel/dashboard.html', context)


# ============================================================================
# MATERIELS
# ============================================================================

@login_required
def liste_materiels(request):
    """Liste du matériel avec filtres."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_gerer_materiel(employe) and not _peut_affecter_materiel(employe):
        messages.error(request, "Vous n'avez pas accès à ce module.")
        return redirect('home')

    form_filtres = FiltresMaterielForm(request.GET)
    materiels = MTMT.objects.select_related('CATEGORIE', 'FOURNISSEUR', 'AFFECTE_A').order_by('-DATE_ACQUISITION')

    if form_filtres.is_valid():
        recherche = form_filtres.cleaned_data.get('recherche')
        categorie = form_filtres.cleaned_data.get('categorie')
        statut = form_filtres.cleaned_data.get('statut')
        etat = form_filtres.cleaned_data.get('etat')
        fournisseur = form_filtres.cleaned_data.get('fournisseur')

        if recherche:
            materiels = materiels.filter(
                Q(CODE_INTERNE__icontains=recherche) |
                Q(DESIGNATION__icontains=recherche) |
                Q(NUMERO_SERIE__icontains=recherche) |
                Q(MARQUE__icontains=recherche) |
                Q(MODELE__icontains=recherche)
            )
        if categorie:
            materiels = materiels.filter(CATEGORIE=categorie)
        if statut:
            materiels = materiels.filter(STATUT=statut)
        if etat:
            materiels = materiels.filter(ETAT=etat)
        if fournisseur:
            materiels = materiels.filter(FOURNISSEUR=fournisseur)

    paginator = Paginator(materiels, 20)
    page = request.GET.get('page', 1)
    materiels_page = paginator.get_page(page)

    context = {
        'materiels': materiels_page,
        'form_filtres': form_filtres,
        'peut_gerer': _peut_gerer_materiel(employe),
        'peut_affecter': _peut_affecter_materiel(employe),
    }
    return render(request, 'materiel/liste_materiels.html', context)


@login_required
def detail_materiel(request, uuid):
    """Détail d'un matériel."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_gerer_materiel(employe) and not _peut_affecter_materiel(employe):
        messages.error(request, "Vous n'avez pas accès à ce module.")
        return redirect('home')

    materiel = get_object_or_404(
        MTMT.objects.select_related('CATEGORIE', 'FOURNISSEUR', 'AFFECTE_A', 'CREATED_BY'),
        uuid=uuid
    )

    # Historique des affectations
    affectations = MTAF.objects.filter(MATERIEL=materiel).select_related(
        'EMPLOYE', 'AFFECTE_PAR', 'RETOUR_PAR'
    ).order_by('-DATE_DEBUT')

    # Historique des mouvements
    mouvements = MTMV.objects.filter(MATERIEL=materiel).select_related(
        'EFFECTUE_PAR'
    ).order_by('-DATE_MOUVEMENT')[:20]

    # Historique des maintenances
    maintenances = MTMA.objects.filter(MATERIEL=materiel).select_related(
        'PRESTATAIRE', 'DEMANDE_PAR'
    ).order_by('-DATE_PLANIFIEE')

    context = {
        'materiel': materiel,
        'affectations': affectations,
        'mouvements': mouvements,
        'maintenances': maintenances,
        'peut_gerer': _peut_gerer_materiel(employe),
        'peut_affecter': _peut_affecter_materiel(employe),
    }
    return render(request, 'materiel/detail_materiel.html', context)


@login_required
def creer_materiel(request):
    """Créer un nouveau matériel."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_gerer_materiel(employe):
        messages.error(request, "Vous n'avez pas les droits pour créer du matériel.")
        return redirect('materiel:liste_materiels')

    if request.method == 'POST':
        form = MTMTForm(request.POST, request.FILES)
        if form.is_valid():
            materiel = form.save(commit=False)
            materiel.CREATED_BY = employe
            materiel.STATUT = 'DISPONIBLE'
            materiel.save()

            # Créer le mouvement d'entrée
            MaterielService.creer_mouvement(
                materiel=materiel,
                type_mouvement='ENTREE',
                date_mouvement=materiel.DATE_ACQUISITION,
                effectue_par=employe,
                motif='Acquisition initiale'
            )

            messages.success(request, f"Matériel {materiel.CODE_INTERNE} créé avec succès.")
            return redirect('materiel:detail_materiel', uuid=materiel.uuid)
    else:
        form = MTMTForm()

    return render(request, 'materiel/form_materiel.html', {
        'form': form,
        'titre': 'Nouveau matériel',
        'action': 'Créer',
    })


@login_required
def modifier_materiel(request, uuid):
    """Modifier un matériel existant."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_gerer_materiel(employe):
        messages.error(request, "Vous n'avez pas les droits pour modifier du matériel.")
        return redirect('materiel:liste_materiels')

    materiel = get_object_or_404(MTMT, uuid=uuid)

    if request.method == 'POST':
        form = MTMTEditForm(request.POST, request.FILES, instance=materiel)
        if form.is_valid():
            form.save()
            messages.success(request, f"Matériel {materiel.CODE_INTERNE} modifié avec succès.")
            return redirect('materiel:detail_materiel', uuid=materiel.uuid)
    else:
        form = MTMTEditForm(instance=materiel)

    return render(request, 'materiel/form_materiel.html', {
        'form': form,
        'materiel': materiel,
        'titre': f'Modifier {materiel.CODE_INTERNE}',
        'action': 'Enregistrer',
    })


@login_required
def affecter_materiel(request, uuid):
    """Affecter un matériel à un employé."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_affecter_materiel(employe):
        messages.error(request, "Vous n'avez pas les droits pour affecter du matériel.")
        return redirect('materiel:liste_materiels')

    materiel = get_object_or_404(MTMT, uuid=uuid)

    if materiel.STATUT not in ['DISPONIBLE']:
        messages.warning(request, f"Ce matériel n'est pas disponible pour affectation (statut: {materiel.get_STATUT_display()}).")
        return redirect('materiel:detail_materiel', uuid=materiel.uuid)

    if request.method == 'POST':
        form = AffectationForm(request.POST)
        if form.is_valid():
            employe_id = form.cleaned_data['employe_id']
            try:
                employe_dest = ZY00.objects.get(matricule=employe_id)
            except ZY00.DoesNotExist:
                messages.error(request, "Employé non trouvé.")
                return redirect('materiel:affecter_materiel', uuid=materiel.uuid)

            try:
                MaterielService.affecter_materiel(
                    materiel=materiel,
                    employe=employe_dest,
                    date_debut=timezone.now().date(),
                    affecte_par=employe,
                    type_affectation=form.cleaned_data['type_affectation'],
                    date_retour_prevue=form.cleaned_data.get('date_retour_prevue'),
                    motif=form.cleaned_data.get('motif')
                )
                messages.success(request, f"Matériel {materiel.CODE_INTERNE} affecté à {employe_dest.nom} {employe_dest.prenoms}.")
                return redirect('materiel:detail_materiel', uuid=materiel.uuid)
            except ValueError as e:
                messages.error(request, str(e))
    else:
        form = AffectationForm()

    return render(request, 'materiel/form_affectation.html', {
        'form': form,
        'materiel': materiel,
    })


@login_required
def retourner_materiel(request, uuid):
    """Retourner un matériel affecté."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_affecter_materiel(employe):
        messages.error(request, "Vous n'avez pas les droits pour gérer les retours.")
        return redirect('materiel:liste_materiels')

    materiel = get_object_or_404(MTMT, uuid=uuid)

    if materiel.STATUT not in ['AFFECTE', 'EN_PRET']:
        messages.warning(request, "Ce matériel n'est pas actuellement affecté.")
        return redirect('materiel:detail_materiel', uuid=materiel.uuid)

    # Récupérer l'affectation active
    affectation = MTAF.objects.filter(
        MATERIEL=materiel,
        ACTIF=True,
        DATE_FIN__isnull=True
    ).first()

    if not affectation:
        messages.warning(request, "Aucune affectation active trouvée pour ce matériel.")
        return redirect('materiel:detail_materiel', uuid=materiel.uuid)

    if request.method == 'POST':
        form = RetourForm(request.POST)
        if form.is_valid():
            try:
                MaterielService.retourner_materiel(
                    affectation=affectation,
                    date_retour=timezone.now().date(),
                    retour_par=employe,
                    etat_retour=form.cleaned_data['etat_retour'],
                    commentaire=form.cleaned_data.get('observations')
                )
                messages.success(request, f"Retour du matériel {materiel.CODE_INTERNE} enregistré.")
                return redirect('materiel:detail_materiel', uuid=materiel.uuid)
            except ValueError as e:
                messages.error(request, str(e))
    else:
        form = RetourForm(initial={'etat_retour': materiel.ETAT})

    return render(request, 'materiel/form_retour.html', {
        'form': form,
        'materiel': materiel,
    })


@login_required
def reformer_materiel(request, uuid):
    """Réformer un matériel."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_gerer_materiel(employe):
        messages.error(request, "Vous n'avez pas les droits pour réformer du matériel.")
        return redirect('materiel:liste_materiels')

    materiel = get_object_or_404(MTMT, uuid=uuid)

    if materiel.STATUT == 'REFORME':
        messages.warning(request, "Ce matériel est déjà réformé.")
        return redirect('materiel:detail_materiel', uuid=materiel.uuid)

    if request.method == 'POST':
        form = ReformeForm(request.POST)
        if form.is_valid():
            try:
                MaterielService.reformer_materiel(
                    materiel=materiel,
                    date_reforme=timezone.now().date(),
                    motif=form.cleaned_data['motif'],
                    effectue_par=employe,
                    valeur_sortie=form.cleaned_data.get('valeur_residuelle')
                )
                messages.success(request, f"Matériel {materiel.CODE_INTERNE} réformé.")
                return redirect('materiel:detail_materiel', uuid=materiel.uuid)
            except ValueError as e:
                messages.error(request, str(e))
    else:
        form = ReformeForm(initial={'valeur_residuelle': materiel.valeur_residuelle})

    return render(request, 'materiel/form_reforme.html', {
        'form': form,
        'materiel': materiel,
    })


# ============================================================================
# MAINTENANCES
# ============================================================================

@login_required
def liste_maintenances(request):
    """Liste des maintenances."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_gerer_materiel(employe):
        messages.error(request, "Vous n'avez pas accès à ce module.")
        return redirect('home')

    form_filtres = FiltresMaintenanceForm(request.GET)
    maintenances = MTMA.objects.select_related(
        'MATERIEL', 'MATERIEL__CATEGORIE', 'PRESTATAIRE', 'DEMANDE_PAR'
    ).order_by('-DATE_PLANIFIEE')

    if form_filtres.is_valid():
        type_maint = form_filtres.cleaned_data.get('type_maintenance')
        statut = form_filtres.cleaned_data.get('statut')
        date_debut = form_filtres.cleaned_data.get('date_debut')
        date_fin = form_filtres.cleaned_data.get('date_fin')

        if type_maint:
            maintenances = maintenances.filter(TYPE_MAINTENANCE=type_maint)
        if statut:
            maintenances = maintenances.filter(STATUT=statut)
        if date_debut:
            maintenances = maintenances.filter(DATE_PLANIFIEE__gte=date_debut)
        if date_fin:
            maintenances = maintenances.filter(DATE_PLANIFIEE__lte=date_fin)

    paginator = Paginator(maintenances, 20)
    page = request.GET.get('page', 1)
    maintenances_page = paginator.get_page(page)

    context = {
        'maintenances': maintenances_page,
        'form_filtres': form_filtres,
    }
    return render(request, 'materiel/liste_maintenances.html', context)


@login_required
def creer_maintenance(request, uuid):
    """Planifier une maintenance pour un matériel."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_gerer_materiel(employe):
        messages.error(request, "Vous n'avez pas les droits pour planifier des maintenances.")
        return redirect('materiel:liste_materiels')

    materiel = get_object_or_404(MTMT, uuid=uuid)

    if request.method == 'POST':
        form = MTMAForm(request.POST)
        if form.is_valid():
            try:
                maintenance = MaterielService.planifier_maintenance(
                    materiel=materiel,
                    type_maintenance=form.cleaned_data['TYPE_MAINTENANCE'],
                    date_planifiee=form.cleaned_data['DATE_PLANIFIEE'],
                    description=form.cleaned_data.get('DESCRIPTION'),
                    demande_par=employe,
                    prestataire=form.cleaned_data.get('PRESTATAIRE'),
                    intervenant_interne=form.cleaned_data.get('INTERVENANT_INTERNE')
                )
                messages.success(request, f"Maintenance {maintenance.REFERENCE} planifiée.")
                return redirect('materiel:detail_materiel', uuid=materiel.uuid)
            except ValueError as e:
                messages.error(request, str(e))
    else:
        form = MTMAForm()

    return render(request, 'materiel/form_maintenance.html', {
        'form': form,
        'materiel': materiel,
        'titre': 'Planifier une maintenance',
    })


@login_required
def demarrer_maintenance(request, pk):
    """Démarrer une maintenance planifiée."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_gerer_materiel(employe):
        return JsonResponse({'success': False, 'message': 'Non autorisé'}, status=403)

    maintenance = get_object_or_404(MTMA, pk=pk)

    try:
        MaterielService.demarrer_maintenance(maintenance)
        messages.success(request, f"Maintenance {maintenance.REFERENCE} démarrée.")
    except ValueError as e:
        messages.error(request, str(e))

    return redirect('materiel:liste_maintenances')


@login_required
def terminer_maintenance(request, pk):
    """Terminer une maintenance en cours."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_gerer_materiel(employe):
        messages.error(request, "Vous n'avez pas les droits pour gérer les maintenances.")
        return redirect('materiel:liste_maintenances')

    maintenance = get_object_or_404(MTMA.objects.select_related('MATERIEL'), pk=pk)

    if maintenance.STATUT not in ['PLANIFIE', 'EN_COURS']:
        messages.warning(request, "Cette maintenance ne peut pas être terminée.")
        return redirect('materiel:liste_maintenances')

    if request.method == 'POST':
        form = TerminerMaintenanceForm(request.POST)
        if form.is_valid():
            try:
                MaterielService.terminer_maintenance(
                    maintenance=maintenance,
                    date_fin=form.cleaned_data['date_fin'],
                    resultat=form.cleaned_data.get('rapport') or '',
                    etat_apres=form.cleaned_data['etat_materiel'],
                    cout_pieces=form.cleaned_data.get('cout_pieces') or Decimal('0'),
                    cout_main_oeuvre=form.cleaned_data.get('cout_main_oeuvre') or Decimal('0')
                )
                messages.success(request, f"Maintenance {maintenance.REFERENCE} terminée.")
                return redirect('materiel:detail_materiel', uuid=maintenance.MATERIEL.uuid)
            except ValueError as e:
                messages.error(request, str(e))
    else:
        form = TerminerMaintenanceForm()

    return render(request, 'materiel/form_terminer_maintenance.html', {
        'form': form,
        'maintenance': maintenance,
    })


# ============================================================================
# CATEGORIES
# ============================================================================

@login_required
def liste_categories(request):
    """Liste des catégories de matériel."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_gerer_materiel(employe):
        messages.error(request, "Vous n'avez pas accès à ce module.")
        return redirect('home')

    categories = MTCA.objects.annotate_stats().order_by('ORDRE', 'LIBELLE')

    return render(request, 'materiel/liste_categories.html', {
        'categories': categories,
    })


@login_required
def creer_categorie(request):
    """Créer une catégorie."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_gerer_materiel(employe):
        return JsonResponse({'success': False, 'message': 'Non autorisé'}, status=403)

    if request.method == 'POST':
        form = MTCAForm(request.POST)
        if form.is_valid():
            form.save()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Catégorie créée'})
            messages.success(request, 'Catégorie créée avec succès.')
            return redirect('materiel:liste_categories')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    else:
        form = MTCAForm()

    return render(request, 'materiel/form_categorie.html', {
        'form': form,
        'titre': 'Nouvelle catégorie',
    })


@login_required
def modifier_categorie(request, pk):
    """Modifier une catégorie."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_gerer_materiel(employe):
        return JsonResponse({'success': False, 'message': 'Non autorisé'}, status=403)

    categorie = get_object_or_404(MTCA, pk=pk)

    if request.method == 'POST':
        form = MTCAForm(request.POST, instance=categorie)
        if form.is_valid():
            form.save()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Catégorie modifiée'})
            messages.success(request, 'Catégorie modifiée avec succès.')
            return redirect('materiel:liste_categories')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    else:
        form = MTCAForm(instance=categorie)

    return render(request, 'materiel/form_categorie.html', {
        'form': form,
        'categorie': categorie,
        'titre': f'Modifier {categorie.LIBELLE}',
    })


# ============================================================================
# FOURNISSEURS
# ============================================================================

@login_required
def liste_fournisseurs(request):
    """Liste des fournisseurs."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_gerer_materiel(employe):
        messages.error(request, "Vous n'avez pas accès à ce module.")
        return redirect('home')

    fournisseurs = MTFO.objects.annotate_stats().order_by('RAISON_SOCIALE')

    return render(request, 'materiel/liste_fournisseurs.html', {
        'fournisseurs': fournisseurs,
    })


@login_required
def creer_fournisseur(request):
    """Créer un fournisseur."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_gerer_materiel(employe):
        return JsonResponse({'success': False, 'message': 'Non autorisé'}, status=403)

    if request.method == 'POST':
        form = MTFOForm(request.POST)
        if form.is_valid():
            form.save()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Fournisseur créé'})
            messages.success(request, 'Fournisseur créé avec succès.')
            return redirect('materiel:liste_fournisseurs')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    else:
        form = MTFOForm()

    return render(request, 'materiel/form_fournisseur.html', {
        'form': form,
        'titre': 'Nouveau fournisseur',
    })


@login_required
def modifier_fournisseur(request, pk):
    """Modifier un fournisseur."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_gerer_materiel(employe):
        return JsonResponse({'success': False, 'message': 'Non autorisé'}, status=403)

    fournisseur = get_object_or_404(MTFO, pk=pk)

    if request.method == 'POST':
        form = MTFOForm(request.POST, instance=fournisseur)
        if form.is_valid():
            form.save()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Fournisseur modifié'})
            messages.success(request, 'Fournisseur modifié avec succès.')
            return redirect('materiel:liste_fournisseurs')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    else:
        form = MTFOForm(instance=fournisseur)

    return render(request, 'materiel/form_fournisseur.html', {
        'form': form,
        'fournisseur': fournisseur,
        'titre': f'Modifier {fournisseur.RAISON_SOCIALE}',
    })


# ============================================================================
# MON MATERIEL (vue employé)
# ============================================================================

@login_required
def mon_materiel(request):
    """Affiche le matériel affecté à l'employé connecté."""
    employe = getattr(request.user, 'employe', None)
    if not employe:
        messages.error(request, "Profil employé non trouvé.")
        return redirect('home')

    stats = StatistiquesMaterielService.get_stats_employe(employe)

    # Historique des affectations
    historique = MTAF.objects.filter(EMPLOYE=employe).select_related(
        'MATERIEL', 'MATERIEL__CATEGORIE', 'AFFECTE_PAR'
    ).order_by('-DATE_DEBUT')[:20]

    context = {
        'stats': stats,
        'historique': historique,
    }
    return render(request, 'materiel/mon_materiel.html', context)


# ============================================================================
# EXPORTS
# ============================================================================

@login_required
def export_materiels_excel(request):
    """Exporter la liste du matériel en Excel."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_gerer_materiel(employe):
        messages.error(request, "Vous n'avez pas les droits pour exporter.")
        return redirect('materiel:liste_materiels')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matériels"

    # Style en-tête
    header_fill = PatternFill(start_color="1c5d5f", end_color="1c5d5f", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        'Code interne', 'Désignation', 'Catégorie', 'Marque', 'Modèle',
        'N° Série', 'Statut', 'État', 'Affecté à', 'Date acquisition',
        'Prix acquisition', 'Localisation'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Données
    materiels = MTMT.objects.select_related('CATEGORIE', 'AFFECTE_A').order_by('CODE_INTERNE')

    for row, mat in enumerate(materiels, 2):
        ws.cell(row=row, column=1, value=mat.CODE_INTERNE).border = thin_border
        ws.cell(row=row, column=2, value=mat.DESIGNATION).border = thin_border
        ws.cell(row=row, column=3, value=mat.CATEGORIE.LIBELLE if mat.CATEGORIE else '').border = thin_border
        ws.cell(row=row, column=4, value=mat.MARQUE or '').border = thin_border
        ws.cell(row=row, column=5, value=mat.MODELE or '').border = thin_border
        ws.cell(row=row, column=6, value=mat.NUMERO_SERIE or '').border = thin_border
        ws.cell(row=row, column=7, value=mat.get_STATUT_display()).border = thin_border
        ws.cell(row=row, column=8, value=mat.get_ETAT_display()).border = thin_border
        ws.cell(row=row, column=9, value=f"{mat.AFFECTE_A.nom} {mat.AFFECTE_A.prenoms}" if mat.AFFECTE_A else '').border = thin_border
        ws.cell(row=row, column=10, value=mat.DATE_ACQUISITION.strftime('%d/%m/%Y') if mat.DATE_ACQUISITION else '').border = thin_border
        ws.cell(row=row, column=11, value=float(mat.PRIX_ACQUISITION) if mat.PRIX_ACQUISITION else 0).border = thin_border
        ws.cell(row=row, column=12, value=mat.LOCALISATION or '').border = thin_border

    # Ajuster largeur colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=materiels_{timezone.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


# ============================================================================
# API (Autocomplete, etc.)
# ============================================================================

@login_required
@require_GET
def api_search_employes(request):
    """Recherche d'employés pour autocomplete."""
    query = request.GET.get('q', '').strip()
    if len(query) < 2:
        return JsonResponse({'results': []})

    employes = ZY00.objects.filter(
        Q(nom__icontains=query) |
        Q(prenoms__icontains=query) |
        Q(matricule__icontains=query)
    ).filter(etat='actif')[:10]

    results = [
        {
            'id': emp.matricule,
            'text': f"{emp.nom} {emp.prenoms} ({emp.matricule})",
            'matricule': emp.matricule,
            'nom': emp.nom,
            'prenoms': emp.prenoms,
        }
        for emp in employes
    ]

    return JsonResponse({'results': results})


@login_required
@require_GET
def api_stats_dashboard(request):
    """API pour les stats du dashboard (refresh AJAX)."""
    employe = getattr(request.user, 'employe', None)
    if not _peut_gerer_materiel(employe):
        return JsonResponse({'error': 'Non autorisé'}, status=403)

    stats = StatistiquesMaterielService.get_stats_globales()
    alertes = StatistiquesMaterielService.get_alertes()

    return JsonResponse({
        'stats': {
            'total_materiels': stats['total_materiels'],
            'disponibles': stats['disponibles'],
            'affectes': stats['affectes'],
            'en_maintenance': stats['en_maintenance'],
            'valeur_totale': float(stats['valeur_totale']),
        },
        'nb_alertes': alertes['nb_alertes'],
    })
